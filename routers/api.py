import os, io, shutil
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Response, Request
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from typing import Optional
from database import get_conn
from auth import hash_password, verify_password, create_token, get_current_user, require_admin, decode_token

router = APIRouter()

UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "/home/claude/ardiyamidly/static/uploads")
VIDEO_DIR  = os.path.join(UPLOAD_DIR, "videos")
PHOTO_DIR  = os.path.join(UPLOAD_DIR, "photos")
os.makedirs(VIDEO_DIR, exist_ok=True)
os.makedirs(PHOTO_DIR, exist_ok=True)

# ─────────────── AUTH ───────────────
class LoginBody(BaseModel):
    nama: str
    sandi: str

@router.post("/api/login")
def login(body: LoginBody, response: Response):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE nama=%s", (body.nama,))
    user = cur.fetchone()
    cur.close(); conn.close()
    if not user or not verify_password(body.sandi, user["sandi_hash"]):
        raise HTTPException(400, "Nama atau sandi salah")
    token = create_token(user["id"], user["nama"], user["role"])
    response.set_cookie("token", token, httponly=True, max_age=86400)
    return {"role": user["role"], "nama": user["nama"]}

@router.post("/api/logout")
def logout(response: Response):
    response.delete_cookie("token")
    return {"ok": True}

@router.get("/api/me")
def me(user=Depends(get_current_user)):
    return user

class RegisterBody(BaseModel):
    nama: str
    sandi: str
    role: str = "viewer"

@router.post("/api/register")
def register(body: RegisterBody):
    """Public register — selalu jadi viewer"""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE nama=%s", (body.nama,))
    if cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(400, "Nama sudah digunakan")
    hashed = hash_password(body.sandi)
    cur.execute(
        "INSERT INTO users (nama, sandi_hash, role) VALUES (%s,%s,'viewer') RETURNING id",
        (body.nama, hashed)
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "nama": body.nama, "role": "viewer"}

@router.post("/api/users")
def add_user(body: RegisterBody, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE nama=%s", (body.nama,))
    if cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(400, "Nama sudah digunakan")
    hashed = hash_password(body.sandi)
    cur.execute("INSERT INTO users (nama, sandi_hash, role) VALUES (%s,%s,%s) RETURNING id",
                (body.nama, hashed, body.role))
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "nama": body.nama, "role": body.role}

@router.get("/api/users")
def list_users(user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT id, nama, role, created_at FROM users ORDER BY id")
    rows = cur.fetchall(); cur.close(); conn.close()
    return rows

@router.delete("/api/users/{uid}")
def delete_user(uid: int, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id=%s AND role!='admin' RETURNING id", (uid,))
    row = cur.fetchone(); conn.commit(); cur.close(); conn.close()
    if not row:
        raise HTTPException(404, "User tidak ditemukan atau tidak bisa dihapus")
    return {"ok": True}

# ─────────────── RESEARCH DATA ───────────────
@router.get("/api/research")
def list_research(user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM research_data ORDER BY id")
    rows = cur.fetchall(); cur.close(); conn.close()
    return rows

@router.post("/api/research")
def add_research(
    nama_ibu: str = Form(...),
    pekerjaan: str = Form(""),
    yang_dirasakan: str = Form(""),
    user=Depends(require_admin)
):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO research_data (nama_ibu, pekerjaan, yang_dirasakan) VALUES (%s,%s,%s) RETURNING id",
        (nama_ibu, pekerjaan, yang_dirasakan)
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id}

@router.put("/api/research/{rid}")
def update_research(
    rid: int,
    nama_ibu: str = Form(...),
    pekerjaan: str = Form(""),
    yang_dirasakan: str = Form(""),
    user=Depends(require_admin)
):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        "UPDATE research_data SET nama_ibu=%s, pekerjaan=%s, yang_dirasakan=%s, updated_at=NOW() WHERE id=%s",
        (nama_ibu, pekerjaan, yang_dirasakan, rid)
    )
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/api/research/{rid}")
def delete_research(rid: int, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM research_data WHERE id=%s", (rid,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

# ─────────────── FEEDBACK ───────────────
@router.get("/api/feedback")
def list_feedback():
    """Publik — semua orang termasuk Viewer & non-login bisa lihat (seperti wall komentar)"""
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM feedback ORDER BY created_at DESC")
    rows = cur.fetchall(); cur.close(); conn.close()
    return rows

@router.post("/api/feedback")
async def add_feedback(request: Request):
    # Viewer & admin boleh submit feedback — cek token optional
    form = await request.form()
    nama_ibu = form.get("nama_ibu","")
    pengalaman = form.get("pengalaman","")
    perasaan_setelah = form.get("perasaan_setelah","")
    if not nama_ibu:
        raise HTTPException(400, "Nama ibu wajib diisi")
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO feedback (nama_ibu, pengalaman, perasaan_setelah) VALUES (%s,%s,%s) RETURNING id",
        (nama_ibu, pengalaman, perasaan_setelah)
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id}

@router.delete("/api/feedback/{fid}")
def delete_feedback(fid: int, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("DELETE FROM feedback WHERE id=%s", (fid,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.get("/api/feedback/export")
def export_feedback(user=Depends(require_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM feedback ORDER BY created_at DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Feedback Ibu"

    headers = ["No", "Nama Ibu", "Pengalaman Menjaga Anak", "Perasaan Setelah Menggunakan Aplikasi", "Tanggal Dikirim"]
    sheet.append(headers)

    header_fill = PatternFill("solid", start_color="6DBF96", end_color="6DBF96")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD")
    )
    for col_idx, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for i, r in enumerate(rows, start=1):
        tanggal = r["created_at"].strftime("%d-%m-%Y %H:%M") if r["created_at"] else "-"
        row_data = [i, r["nama_ibu"], r["pengalaman"] or "-", r["perasaan_setelah"] or "-", tanggal]
        sheet.append(row_data)
        row_idx = i + 1
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['C'].width = 45
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['E'].width = 18
    sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"feedback_ardiamindly_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# ─────────────── GALLERY ───────────────
@router.get("/api/gallery")
def list_gallery(user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM gallery_photos ORDER BY created_at DESC")
    rows = cur.fetchall(); cur.close(); conn.close()
    return rows

@router.post("/api/gallery")
async def upload_photo(
    foto: UploadFile = File(...),
    deskripsi: str = Form(""),
    user=Depends(require_admin)
):
    import uuid, aiofiles
    ext = foto.filename.rsplit(".", 1)[-1].lower()
    if ext not in ["jpg","jpeg","png","gif","webp","pdf"]:
        raise HTTPException(400, "Format tidak didukung")
    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = os.path.join(PHOTO_DIR, fname)
    async with aiofiles.open(fpath, "wb") as f:
        while chunk := await foto.read(512 * 1024):  # 512KB chunks
            await f.write(chunk)
    conn = get_conn(); cur = conn.cursor()
    cur.execute(
        "INSERT INTO gallery_photos (filename, deskripsi, uploaded_by) VALUES (%s,%s,%s) RETURNING id",
        (fname, deskripsi, user["nama"])
    )
    new_id = cur.fetchone()["id"]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "filename": fname}

@router.put("/api/gallery/{gid}")
def update_gallery(
    gid: int,
    deskripsi: str = Form(""),
    user=Depends(require_admin)
):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("UPDATE gallery_photos SET deskripsi=%s WHERE id=%s", (deskripsi, gid))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@router.delete("/api/gallery/{gid}")
def delete_gallery(gid: int, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT filename FROM gallery_photos WHERE id=%s", (gid,))
    row = cur.fetchone()
    if row:
        fpath = os.path.join(PHOTO_DIR, row["filename"])
        if os.path.exists(fpath):
            os.remove(fpath)
        cur.execute("DELETE FROM gallery_photos WHERE id=%s", (gid,))
        conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@router.get("/uploads/photos/{filename}")
def serve_photo(filename: str):
    fpath = os.path.join(PHOTO_DIR, filename)
    if not os.path.exists(fpath):
        raise HTTPException(404, "File tidak ditemukan")
    return FileResponse(fpath)

# ─────────────── VIDEO ───────────────
@router.get("/api/videos")
def list_videos(user=Depends(get_current_user)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM videos ORDER BY fitur_id, urutan")
    rows = cur.fetchall(); cur.close(); conn.close()
    return rows

@router.post("/api/videos")
async def upload_video(
    fitur_id: int = Form(...),
    judul: str = Form(...),
    judul_en: str = Form(""),
    durasi: str = Form(""),
    urutan: int = Form(1),
    video: UploadFile = File(None),
    youtube_url: str = Form(""),
    user=Depends(require_admin)
):
    import uuid, aiofiles, re

    youtube_url = youtube_url.strip()

    if youtube_url:
        # ── Mode YouTube embed ──
        patterns = [
            r"(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/|youtube\.com\/shorts\/)([A-Za-z0-9_-]{11})"
        ]
        yt_id = None
        for p in patterns:
            m = re.search(p, youtube_url)
            if m:
                yt_id = m.group(1)
                break
        if not yt_id:
            raise HTTPException(400, "Link YouTube tidak valid")

        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            "INSERT INTO videos (fitur_id, judul, judul_en, durasi, filename, source_type, youtube_id, urutan) "
            "VALUES (%s,%s,%s,%s,NULL,'youtube',%s,%s) RETURNING id",
            (fitur_id, judul, judul_en, durasi, yt_id, urutan)
        )
        new_id = cur.fetchone()["id"]
        conn.commit(); cur.close(); conn.close()
        return {"id": new_id, "youtube_id": yt_id}

    elif video is not None and video.filename:
        # ── Mode upload file ──
        ext = video.filename.rsplit(".",1)[-1].lower()
        if ext not in ["mp4","webm","mov"]:
            raise HTTPException(400, "Format video tidak didukung (mp4/webm/mov)")
        fname = f"{uuid.uuid4().hex}.{ext}"
        fpath = os.path.join(VIDEO_DIR, fname)
        async with aiofiles.open(fpath, "wb") as f:
            while chunk := await video.read(1024 * 1024):  # 1MB chunks
                await f.write(chunk)
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            "INSERT INTO videos (fitur_id, judul, judul_en, durasi, filename, source_type, urutan) "
            "VALUES (%s,%s,%s,%s,%s,'upload',%s) RETURNING id",
            (fitur_id, judul, judul_en, durasi, fname, urutan)
        )
        new_id = cur.fetchone()["id"]
        conn.commit(); cur.close(); conn.close()
        return {"id": new_id, "filename": fname}

    else:
        raise HTTPException(400, "Pilih file video atau masukkan link YouTube")

@router.delete("/api/videos/{vid}")
def delete_video(vid: int, user=Depends(require_admin)):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT filename FROM videos WHERE id=%s", (vid,))
    row = cur.fetchone()
    if row:
        if row["filename"]:
            fpath = os.path.join(VIDEO_DIR, row["filename"])
            if os.path.exists(fpath):
                os.remove(fpath)
        cur.execute("DELETE FROM videos WHERE id=%s", (vid,))
        conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@router.get("/stream/video/{filename}")
def stream_video(filename: str, request: Request, user=Depends(get_current_user)):
    fpath = os.path.join(VIDEO_DIR, filename)
    if not os.path.exists(fpath):
        raise HTTPException(404, "Video tidak ditemukan")
    file_size = os.path.getsize(fpath)
    range_header = request.headers.get("range")
    chunk = 1024 * 1024  # 1MB chunks

    if range_header:
        start_str, end_str = range_header.replace("bytes=", "").split("-")
        start = int(start_str)
        end = int(end_str) if end_str else min(start + chunk - 1, file_size - 1)
        length = end - start + 1

        def gen():
            with open(fpath, "rb") as f:
                f.seek(start)
                remaining = length
                while remaining:
                    data = f.read(min(8192, remaining))
                    if not data:
                        break
                    remaining -= len(data)
                    yield data

        return StreamingResponse(
            gen(),
            status_code=206,
            media_type="video/mp4",
            headers={
                "Content-Range": f"bytes {start}-{end}/{file_size}",
                "Accept-Ranges": "bytes",
                "Content-Length": str(length),
            }
        )

    def gen_full():
        with open(fpath, "rb") as f:
            while True:
                data = f.read(8192)
                if not data:
                    break
                yield data

    return StreamingResponse(
        gen_full(),
        media_type="video/mp4",
        headers={
            "Content-Length": str(file_size),
            "Accept-Ranges": "bytes",
        }
    )

# ─────────────── QR CODE ───────────────
@router.get("/api/qr")
def generate_qr(request: Request, user=Depends(get_current_user)):
    import qrcode
    base_url = str(request.base_url).rstrip("/")
    img = qrcode.make(base_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")